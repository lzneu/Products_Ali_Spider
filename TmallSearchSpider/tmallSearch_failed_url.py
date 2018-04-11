# -*- coding: utf-8 -*-
"""
Created on Mon Feb 20 09:27:42 2017

@author: Yang
"""
import random
import time

# firefox_capabilities = DesiredCapabilities.FIREFOX
# # firefox_capabilities['marionette'] = True
# driver = webdriver.Firefox(capabilities=firefox_capabilities,executable_path='F:/geckodriver')
# wait = ui.WebDriverWait(driver,10)
from bs4 import BeautifulSoup
from openpyxl.cell import *
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
# import pymysql
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException, \
    ElementNotInteractableException
from selenium.webdriver.support.wait import WebDriverWait

from Config.Tmall_Products_Config import urls_collections_config
from Config.config import *
from MongoRelated.mongoConnection import connect_mongo
# 连接数据库
# db = pymysql.connect(host = 'localhost',port = 3306,user = 'root',passwd = '123456',db = 'jd_data',charset = 'utf8')
from Tools import loginTmall
from Tools.check_CN import check_contain_chinese
from Tools.logger import Logger


def insertToMongo(mongo_collection, oneDict):
    cursor = mongo_collection.find({'original_id': oneDict['original_id']})
    # print(oneDict)
    res = list(cursor)
    if len(res) == 0:
        # timeStr =oneDict['creationTime']     # 时间转换
        # oneDict['creationTime'] =convertTimeStringToDateTime(timeStr)
        mongo_collection.insert(oneDict)
        # print(res)
    else:
        print(oneDict['original_id'], '已存在')
        key = oneDict['Utime']
        res[0]['price'][key] = oneDict['price'][key]
        res[0]['Utime'] = oneDict['Utime']
        res[0]['sell_num'][key] = oneDict['sell_num'][key]
        res[0]['comment_quantity'][key] = oneDict['comment_quantity'][key]
        mongo_collection.update({"original_id": oneDict['original_id']}, {"$set": res[0]})


def get_shop_type(seller_name):
    if '天猫超市' == seller_name:
        return 1
    elif 'costco海外旗舰店' == seller_name or '苏宁易购官方旗舰店' == seller_name or '我买网官方旗舰店' == seller_name:
        return 2
    elif '旗舰店' in seller_name:
        return 3
    else:
        return 4


def dealWith(mongo_collection, bs_obj, key, brand):
    # 初始化各个变量
    category_id = brand['category_id']
    brand_original_name = brand['original_name']
    brand_id = brand['id']
    brand_original_id = brand['original_id']
    brand_original_name_en = brand['original_name_en']
    brand_original_name_cn = brand['original_name_cn']
    brand_name = brand['name']
    brand_en_name = brand['en_name']
    category = brand['category']
    store_id = brand['store_id']


    items = bs_obj.find_all('div', {'class': 'product'})
    num = len(items)
    for item in items:
        try:
            price = item.find('p', {'class': 'productPrice'}).find('em').get('title')
            print(price)
            try:
                product_title = item.find('p', {'class': 'productTitle'}).find('a').get('title')
            except Exception as e:
                try:
                    product_title = item.find('div', {'class': 'productTitle productTitle-spu'}).find('a').text.strip()
                except:
                    product_title = item.find('div', {'class': 'productTitle'}).find('a').get('title')

            href = 'http:' + item.find('div', {'class': 'productImg-wrap'}).find('a').get('href')

            print(product_title)
            seller_name = item.find('div', {'class': 'productShop'}).find('a').text.strip()
            print(seller_name)
            seller_href = 'http:' + item.find('div', {'class': 'productShop'}).find('a').get('href')
            print(seller_href)
            # //store.taobao.com/search.htm?spm=a220m.1000858.1000725.3.4c6378104NCGZR&user_number_id=916243692&rn=8f2eeb3a7f2b30d93e1dd40721e39e22&keyword=洗衣液
            seller_id = seller_href.split('user_number_id=')[1].split('&rn=')[0]
            print(seller_id)
            product_id = item.get('data-id')
            print(product_id)
            sell_num = item.find('p', {'class': 'productStatus'}).find_all('span')[0].text
            print(sell_num)
            Utime = time.strftime("%d/%m/%Y")
            try:
                commits = item.find('p', {'class': 'productStatus'}).find_all('span')[1].text
            except:
                commits = ''
            print(commits)

            price_dic = {}
            sell_num_dic = {}
            commits_dic = {}

            price_dic[Utime] = price
            sell_num_dic[Utime] = sell_num
            commits_dic[Utime] = commits

            shop_type = get_shop_type(seller_name)

            item_dict = {}
            item_dict['price'] = price_dic
            item_dict['shop_id'] = int(seller_id)
            item_dict['href'] = href
            item_dict['name'] = product_title
            item_dict['shop_name'] = seller_name
            item_dict['shop_href'] = seller_href
            item_dict['original_id'] = product_id
            item_dict['sell_num'] = sell_num_dic
            item_dict['category'] = category
            item_dict['Utime'] = Utime
            item_dict['comment_quantity'] = commits_dic
            item_dict['brand_original_name'] = brand_original_name
            item_dict['brand_id'] = brand_id
            item_dict['brand_original_id'] = brand_original_id
            item_dict['brand_name'] = brand_name
            item_dict['brand_en_name'] = brand_en_name
            item_dict['brand_original_name_en'] = brand_original_name_en
            item_dict['brand_original_name_cn'] = brand_original_name_cn
            item_dict['shop_type'] = shop_type
            item_dict['store_id'] = store_id
            item_dict['category_id'] = category_id
            item_dict['status'] = 0
            insertToMongo(mongo_collection, item_dict)
        except Exception as e:
            print('eRROR:')
            print(e.args)
    return num


def getMaxPage(driver):
    try:
        max_page_str = driver.find_element_by_xpath('//*[@id="J_Filter"]/p/b[1]').text
        max_page = max_page_str.split('/')[1]
    except NoSuchElementException:
        return 0
    return max_page


def getCurrentPage(driver):
    current_page_str = driver.find_element_by_xpath('//*[@id="J_Filter"]/p/b[1]').text
    current_page = current_page_str.split('/')[0]
    return current_page


def isNoItem(driver):
    time.sleep(random.uniform(1, 2))

    try:
        driver.find_element_by_class_name('nrt')
    except NoSuchElementException:
        return False
    return True
output = open('fail_url_failed.txt', 'w', encoding='utf-8')
def writeToCsv(url, brand, collection_name):
    output.write(collection_name + '/*/' + str(brand) + '/*/' + url + '\n')


def VisitGoodsPage(mongo_collection, driver, key, brand):
    # 初始化各个变量
    url = brand['original_url']

    try:
        driver.get(url)
    except WebDriverException as e:
        time.sleep(10)
        Logger.info('Error!' + str(e))
        driver.quit()
        driver = loginTmall.login_tmall()
        driver.get(url)
        time.sleep(random.uniform(2, 4))

    time.sleep(random.uniform(0.5, 1))
    # 判断是否跳入了验证码
    current_url = driver.current_url
    if 'https://sec.taobao.com' in current_url:
        collection_name = mongo_collection.name
        writeToCsv(url, brand, collection_name)
        time.sleep(random.uniform(2, 4))
        return
    # 判断是否没有商品
    if isNoItem(driver):
        return
    max_page = getMaxPage(driver)
    if max_page == 0:
        return
    Logger.info('最大页数：' + str(max_page))
    print('准备访问商品页面')
    print('商品详细信息')

    time.sleep(random.uniform(2, 4))
    driver.execute_script("scrollTo(0,1000)")
    time.sleep(random.uniform(1, 2))
    driver.execute_script("scrollTo(0,5000)")
    time.sleep(random.uniform(1, 2))
    driver.execute_script("scrollTo(0,10000)")
    time.sleep(random.uniform(1, 2))
    # driver.execute_script("scrollTo(0,30000)")

    bs_obj = BeautifulSoup(driver.page_source, 'lxml')
    dealWith(mongo_collection, bs_obj, key, brand)
    N = 2
    while N <= int(max_page):

        time.sleep(2)

        element = WebDriverWait(driver, 60).until(lambda driver:
                                                  driver.find_element_by_xpath("//a[@class='ui-page-s-next']"))
        element.click()
        time.sleep(2)
        driver.execute_script("scrollTo(0,1000)")
        time.sleep(1)
        driver.execute_script("scrollTo(0,5000)")
        time.sleep(1)
        driver.execute_script("scrollTo(0,10000)")
        time.sleep(1)
        # driver.execute_script("scrollTo(0,30000)")
        Logger.info(driver.current_url)

        # driver.execute_script("scrollTo(0,30000)")

        bs_obj = BeautifulSoup(driver.page_source, 'lxml')
        dealWith(mongo_collection, bs_obj, key, brand)
        # time.sleep(5)
        current_page = getCurrentPage(driver)
        Logger.info('完成当前页爬取：' + str(current_page))
        if int(current_page) == int(max_page):
            Logger.info(brand['original_name'])
            Logger.info('''
                                    #########################################################################
                                    |                            最大页数爬取完毕                               |
                                    #########################################################################
                                ''')
        N = int(current_page) + 1

    print("done..")


def writeToExcel(dict):
    # inwb = load_workbook(filename="D:/test/洗涤.xlsx")

    wb = Workbook()
    ew = ExcelWriter(workbook=wb)
    dest_filename = '1122.xlsx'
    ws = wb.worksheets[0]
    ws.title = "洗涤产品详情"
    col_list = ['A', 'B', 'C', 'D']
    i = 0
    for key in dict:
        j = 0
        i += 1
        if i <= len(dict):
            for col in col_list:
                data_list = dict[key]
                if j < len(data_list):
                    ws.cell('%s%s' % (col, i)).value = '%s' % (dict[key][j])
                j += 1
    ew.save(filename=dest_filename)

def get_store_id(category):
    for key in store_id_dic.keys():
        if key in category:
            return store_id_dic[key]

    return ''


def updateBrandCollection(category, category_id, brand, mongo_conn, brand_collection_name):
    dataBase = mongo_conn['power']
    collection = dataBase[brand_collection_name]
    dic = {}
    dic['original_name'] = brand[0].strip()
    dic['original_url'] = brand[1]
    dic['original_id'] = brand[1].split('brand=')[1].split('&')[0]
    # 创建category_id 和 store_id
    dic['category_id'] = category_id
    # 天猫-洗衣类--洗衣皂
    dic['store_id'] = get_store_id(category)
    # name en_name  id 初始化为''
    dic['name'] = ''
    dic['en_name'] = ''
    dic['id'] = ''
    dic['status'] = 0
    dic['category'] = category
    original_name_cn = ''
    original_name_en = ''
    brand_original_name_list = dic['original_name'].split('/')
    for word in brand_original_name_list:
        # 中文
        if check_contain_chinese(word):
            original_name_cn = word
        else:
            original_name_en = word

    dic['original_name_en'] = original_name_en
    dic['original_name_cn'] = original_name_cn

    cursor = collection.find({'original_id': dic['original_id']}, {'category': category})

    res = list(cursor)
    if len(res) == 0:
        collection.insert(dic)
        print(dic['original_name'], '已插入')
        # print(res)
    else:
        print(dic['original_name'], '已存在')
    cursor.close()


def setBrandList(category, category_id, driver, url, mongo_conn, brand_collection_name):
    time.sleep(random.uniform(2, 6))
    driver.get(url)
    time.sleep(random.uniform(5, 8))
    brand_list = []
    # //*[@id="J_NavCommonRowItems_0"]/a[1]
    # //*[@id="J_NavAttrsForm"]/div/div[1]/div/div[2]/ul/li[1]/a
    # //*[@id="J_NavAttrsForm"]/div/div[1]/div/div[2]/ul/li[1]/a
    # //*[@id="J_NavAttrsForm"]/div/div[1]/div/div[2]/ul

    # 点击功更过品牌按钮
    # //*[@id="J_NavAttrsForm"]/div/div[1]/div/div[2]/div[2]/a[2]
    try:
        element = WebDriverWait(driver, 60).until(lambda driver:
                                                  driver.find_element_by_xpath(
                                                      '//*[@id="J_NavAttrsForm"]/div/div[1]/div/div[2]/div[2]/a[2]'))
        element.click()
    except ElementNotInteractableException:
        print('天猫美妆')

    time.sleep(3)

    brandElement_list = driver.find_elements_by_xpath('//*[@id="J_NavAttrsForm"]/div/div[1]/div/div[2]/ul/li')

    for brandElement in brandElement_list:
        brand = []
        brand.append(brandElement.find_element_by_tag_name('a').get_attribute('title'))
        brand.append(brandElement.find_element_by_tag_name('a').get_attribute('href'))
        brand_list.append(brand)
    # 更新品牌集合
    for brand in brand_list:
        updateBrandCollection(category, category_id, brand, mongo_conn, brand_collection_name)

    return brand_list


def getBrandList(collection_name, mongo_conn):
    dataBase = mongo_conn['power']
    collection = dataBase[collection_name + '_brands']
    cursors = collection.find()

    res = list(cursors)
    cursors.close()
    return res


if __name__ == '__main__':

    driver = loginTmall.login_tmall()
    time.sleep(1)
    mongo_conn = connect_mongo(mongodb_host, mongodb_port, mongodb_username, mongodb_password)
    dataBase = mongo_conn['power']  # Database

    url_list = []
    with open('fail_url.txt', 'r', encoding='utf-8') as fc:
        for i in fc:
            temp = i.split('/*/')
            url_list.append(temp)
        for url_item in url_list:
            collection_name = url_item[0].strip()
            collection = dataBase[collection_name]
            brand_str = url_item[1].replace('ObjectId', '')
            brand = eval(brand_str)

            Logger.info(brand['original_name'])
            Logger.info(brand['original_url'])
            VisitGoodsPage(collection, driver, brand['category'], brand)
            time.sleep(random.uniform(5, 20))
    driver.quit()
    mongo_conn.close()
    output.close()
